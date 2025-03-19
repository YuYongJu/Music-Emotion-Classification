from google.cloud import storage, videointelligence
from google.oauth2 import service_account
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Get credentials file path from environment variable
current_dir = os.path.dirname(os.path.abspath(__file__))
key_path = os.path.join(current_dir, os.getenv('GOOGLE_APPLICATION_CREDENTIALS'))

# Create a credentials object
credentials = service_account.Credentials.from_service_account_file(key_path)

# Use the credentials when creating the clients
storage_client = storage.Client(credentials=credentials)
video_client = videointelligence.VideoIntelligenceServiceClient(credentials=credentials)

def analyze_videos_in_bucket(bucket_name):
    """Analyze videos in the bucket and return the output file path"""
    try:
        # Use the global video_client that already has credentials
        global video_client, storage_client

        features = [
            videointelligence.Feature.LABEL_DETECTION,
            videointelligence.Feature.SHOT_CHANGE_DETECTION,
            videointelligence.Feature.EXPLICIT_CONTENT_DETECTION
        ]

        # Get the list of objects in the bucket
        # Use global storage_client with credentials
        bucket = storage_client.get_bucket(bucket_name)
        blobs = bucket.list_blobs()

        # Create empty lists to store the results
        label_results = []
        explicit_results = []
        shot_results = []

        video_count = 0
        for blob in blobs:
            if blob.name.endswith('.mp4'):
                video_count += 1
                input_uri = "gs://{}/{}".format(bucket_name, blob.name)
                operation = video_client.annotate_video(
                    request={
                        "features": features,
                        "input_uri": input_uri,
                    }
                )
                print("\nProcessing video {} for label annotations:".format(blob.name))

                result = operation.result(timeout=180)
                print("Finished processing video {}.".format(blob.name))

                # Label detection
                segment_labels = result.annotation_results[0].segment_label_annotations
                for segment_label in segment_labels:
                    label_description = segment_label.entity.description
                    for category_entity in segment_label.category_entities:
                        category_description = category_entity.description

                        for segment in segment_label.segments:
                            start_time = segment.segment.start_time_offset.seconds + segment.segment.start_time_offset.microseconds / 1e6
                            end_time = segment.segment.end_time_offset.seconds + segment.segment.end_time_offset.microseconds / 1e6
                            confidence = segment.confidence

                            # Append the results to the label results list
                            label_results.append({
                                "Video": blob.name,
                                "Label Description": label_description,
                                "Category Description": category_description,
                                "Start Time": start_time,
                                "End Time": end_time,
                                "Confidence": confidence
                            })

                # Explicit content detection
                explicit_content = result.annotation_results[0].explicit_annotation
                video_confidences = []  # List to store confidences for each video
                for frame in explicit_content.frames:
                    time_offset = frame.time_offset.seconds + frame.time_offset.microseconds / 1e6
                    pornography_likelihood = frame.pornography_likelihood

                    # Append the results to the explicit content results list
                    explicit_results.append({
                        "Video": blob.name,
                        "Label Description": "Explicit Content",
                        "Category Description": "N/A",
                        "Start Time": time_offset,
                        "End Time": time_offset,
                        "Confidence": pornography_likelihood
                    })

                    # Append the confidence to the video_confidences list
                    video_confidences.append(pornography_likelihood)

                # Calculate the mean confidence for the video
                video_mean_confidence = sum(video_confidences) / len(video_confidences) if video_confidences else 0

                # Append the mean confidence to the explicit content results for the video
                explicit_results.append({
                    "Video": blob.name,
                    "Label Description": "Mean Confidence",
                    "Category Description": "N/A",
                    "Start Time": None,
                    "End Time": None,
                    "Confidence": video_mean_confidence
                })

                # Shot change detection
                shot_annotations = result.annotation_results[0].shot_annotations
                for shot in shot_annotations:
                    start_time = shot.start_time_offset.seconds + shot.start_time_offset.microseconds / 1e6
                    end_time = shot.end_time_offset.seconds + shot.end_time_offset.microseconds / 1e6

                    # Append the results to the shot detection results list
                    shot_results.append({
                        "Video": blob.name,
                        "Label Description": "Shot Change",
                        "Category Description": "N/A",
                        "Start Time": start_time,
                        "End Time": end_time,
                        "Confidence": None  # No confidence value for shot change detection
                    })

        if video_count == 0:
            print("No videos found in the bucket.")
            return None

        # Create DataFrames from the results
        label_df = pd.DataFrame(label_results)
        explicit_df = pd.DataFrame(explicit_results)
        shot_df = pd.DataFrame(shot_results)

        # Create an Excel workbook
        workbook = Workbook()

        # Create sheets for each table
        label_sheet = workbook.active
        label_sheet.title = "Label Detection"
        explicit_sheet = workbook.create_sheet(title="Explicit Content Detection")
        shot_sheet = workbook.create_sheet(title="Shot Detection")

        # Write tables to the respective sheets
        for row in dataframe_to_rows(label_df, index=False, header=True):
            label_sheet.append(row)

        for row in dataframe_to_rows(explicit_df, index=False, header=True):
            explicit_sheet.append(row)

        for row in dataframe_to_rows(shot_df, index=False, header=True):
            shot_sheet.append(row)

        # Apply formatting to the tables
        header_font = Font(bold=True)
        alignment = Alignment(horizontal="center", vertical="center")

        for cell in label_sheet[1]:
            cell.font = header_font
            cell.alignment = alignment

        for cell in explicit_sheet[1]:
            cell.font = header_font
            cell.alignment = alignment

        for cell in shot_sheet[1]:
            cell.font = header_font
            cell.alignment = alignment

        # Save the Excel file
        excel_file = "GoogleVideoIntelligenceLabelAnalyzer_results.xlsx"
        workbook.save(excel_file)
        print("Results saved to {}".format(excel_file))
        
        return excel_file
    except Exception as e:
        print(f"Error analyzing videos: {e}")
        raise

def list_videos_in_bucket(bucket_name):
    try:
        bucket = storage_client.get_bucket(bucket_name)
        blobs = list(bucket.list_blobs())
        
        video_files = [blob.name for blob in blobs if blob.name.lower().endswith(('.mp4', '.mov', '.avi'))]
        
        if video_files:
            print(f"Found {len(video_files)} videos in bucket '{bucket_name}':")
            for video in video_files:
                print(f"- {video}")
            return True
        else:
            print(f"No videos found in bucket '{bucket_name}'.")
            return False
    except Exception as e:
        print(f"Error accessing bucket '{bucket_name}': {e}")
        return False

def create_bucket_if_not_exists(bucket_name):
    try:
        storage_client.get_bucket(bucket_name)
        print(f"Bucket '{bucket_name}' already exists.")
        return True
    except Exception:
        try:
            bucket = storage_client.create_bucket(bucket_name)
            print(f"Bucket '{bucket_name}' created successfully.")
            return True
        except Exception as e:
            print(f"Failed to create bucket '{bucket_name}': {e}")
            return False

def upload_video_to_bucket(bucket_name, source_file_path, destination_blob_name=None):
    """Uploads a file to the bucket."""
    if destination_blob_name is None:
        destination_blob_name = os.path.basename(source_file_path)
    
    try:
        bucket = storage_client.get_bucket(bucket_name)
        blob = bucket.blob(destination_blob_name)
        
        # Upload the file
        blob.upload_from_filename(source_file_path)
        
        print(f"File {source_file_path} uploaded to {destination_blob_name} in bucket {bucket_name}.")
        return True
    except Exception as e:
        print(f"Error uploading file to bucket: {e}")
        return False

if __name__ == "__main__":
    # Specify your bucket name
    bucket_name = "anime_food_landscape_object_bucket"
    
    # Check if bucket exists and create it if needed
    if create_bucket_if_not_exists(bucket_name):
        # Check if there are videos in the bucket
        has_videos = list_videos_in_bucket(bucket_name)
        
        if has_videos:
            # Call the function to analyze videos in the bucket and save the results as an Excel file
            analyze_videos_in_bucket(bucket_name)
        else:
            print("\nPlease upload videos to the bucket before running analysis.")
            print("You can use the upload_video_to_bucket function to upload videos.")
            print("Example: upload_video_to_bucket(bucket_name, 'path/to/your/video.mp4')")
    else:
        print("Cannot proceed with video analysis without a valid bucket.")

# """Analyze Labels"""
# from google.cloud import videointelligence

# video_client = videointelligence.VideoIntelligenceServiceClient()
# features = [videointelligence.Feature.LABEL_DETECTION]
# operation = video_client.annotate_video(
#     request={
#         "features": features,
#         "input_uri": "gs://anime_food_landscape_object_bucket/Relaxing Anime Cooking ｜ Aesthetic Anime ASMR.mp4",
#     }
# )
# print("\nProcessing video for label annotations:")

# result = operation.result(timeout=180)
# print("\nFinished processing.")

# # first result is retrieved because a single video was processed
# segment_labels = result.annotation_results[0].segment_label_annotations
# for i, segment_label in enumerate(segment_labels):
#     print("Video label description: {}".format(segment_label.entity.description))
#     for category_entity in segment_label.category_entities:
#         print(
#             "\tLabel category description: {}".format(category_entity.description)
#         )

#     for i, segment in enumerate(segment_label.segments):
#         start_time = (
#             segment.segment.start_time_offset.seconds
#             + segment.segment.start_time_offset.microseconds / 1e6
#         )
#         end_time = (
#             segment.segment.end_time_offset.seconds
#             + segment.segment.end_time_offset.microseconds / 1e6
#         )
#         positions = "{}s to {}s".format(start_time, end_time)
#         confidence = segment.confidence
#         print("\tSegment {}: {}".format(i, positions))
#         print("\tConfidence: {}".format(confidence))
#     print("\n")
